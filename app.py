from datetime import date, time, timedelta, datetime
import io
import os
import re
import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.utils.datetime import from_excel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="TRIZ 영업실 업무 프로그램", page_icon="📊", layout="wide", initial_sidebar_state="expanded")

st.markdown('''
<style>
html, body, [class*="css"]{font-family:-apple-system,BlinkMacSystemFont,"Apple SD Gothic Neo","Noto Sans KR",sans-serif;}
.stApp{background:#f6f8fb}.block-container{padding-top:1.4rem;max-width:1380px}section[data-testid="stSidebar"]{background:#fff;border-right:1px solid #e5e7eb}.main-title{font-size:1.75rem;font-weight:900;letter-spacing:-.04em;color:#111827}.subtle{color:#64748b;font-size:.94rem;margin-bottom:1.2rem}[data-testid="stMetric"]{background:#fff;border:1px solid #e5e7eb;border-radius:20px;padding:18px;box-shadow:0 12px 28px rgba(15,23,42,.055);min-height:118px;display:flex;flex-direction:column;justify-content:center;}[data-testid="stMetricValue"]{font-size:1.7rem;font-weight:900;letter-spacing:-.05em}.card{background:#fff;border:1px solid #e5e7eb;border-radius:22px;padding:20px;box-shadow:0 12px 28px rgba(15,23,42,.055);margin-bottom:14px}.dark-card{background:linear-gradient(135deg,#111827,#334155);color:white;border-radius:22px;padding:20px;box-shadow:0 12px 28px rgba(15,23,42,.12);margin-bottom:14px}.dark-card span{color:#cbd5e1}.section-title{font-size:1.45rem;font-weight:900;letter-spacing:-.04em;margin:1rem 0 .35rem}.help{color:#64748b;font-size:.9rem;margin-bottom:.8rem}.chip{display:inline-block;padding:7px 10px;border-radius:999px;font-size:12px;font-weight:800;margin-right:6px;margin-bottom:6px;background:#f8fafc;border:1px solid #e5e7eb;color:#475569}.chip.active{background:#111827;color:white;border-color:#111827}.status-green{display:inline-block;padding:5px 8px;border-radius:999px;background:#dcfce7;color:#166534;font-size:11px;font-weight:900}.calendar-grid{display:grid;grid-template-columns:repeat(5,minmax(140px,1fr));gap:12px}.day-card{min-height:165px;background:#f8fafc;border:1px solid #e5e7eb;border-radius:18px;padding:14px}.day-top{display:flex;justify-content:space-between;align-items:center;color:#64748b;font-size:12px;margin-bottom:10px}.day-date{color:#111827;font-weight:900;font-size:17px}.schedule-item{border-radius:14px;padding:10px;background:#eff6ff;color:#1d4ed8;font-size:12px;line-height:1.38;font-weight:800;margin-top:8px}.schedule-purple{background:#f5f3ff;color:#6d28d9}.schedule-cancel{background:#fff1f2;color:#be123c;text-decoration:line-through;border:1px solid #ffe4e6}.step-box{background:#f8fafc;border:1px solid #e5e7eb;border-radius:18px;padding:16px;height:100%}.step-box b{display:block;font-size:14px;margin-bottom:5px}.step-box span{color:#64748b;font-size:12px}.table-like{border:1px solid #e5e7eb;border-radius:18px;overflow:hidden;background:white}.row{display:grid;grid-template-columns:1.2fr 1fr repeat(4,.8fr);border-bottom:1px solid #eef2f7}.row:last-child{border-bottom:none}.cell{padding:12px 13px;font-size:13px;border-right:1px solid #eef2f7}.cell:last-child{border-right:none}.head .cell{background:#f8fafc;color:#475569;font-weight:900}.group-cell{background:#ecfdf5;color:#047857;font-weight:900;display:flex;align-items:center}.sum .cell{background:#eff6ff;color:#1e40af;font-weight:900}.center{text-align:center}.bar-bg{height:10px;background:#e5e7eb;border-radius:999px;overflow:hidden}.bar-fill{height:100%;background:linear-gradient(90deg,#2563eb,#7c3aed);border-radius:999px}.stButton button,.stDownloadButton button{border-radius:13px;font-weight:800;border:1px solid #111827;background:#111827;color:white}@media(max-width:1000px){.calendar-grid{grid-template-columns:1fr}.row{grid-template-columns:1fr}.cell{border-right:none}}
.field-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:10px 16px;margin-top:12px}
.field-item{background:#f8fafc;border:1px solid #e5e7eb;border-radius:12px;padding:10px 13px}
.field-item b{display:block;font-size:11px;color:#64748b;font-weight:800;margin-bottom:3px}
.field-item span{font-size:13.5px;color:#111827;font-weight:700;white-space:pre-line}
.field-item.full{grid-column:1/-1}
.tag-vendor{background:#eff6ff !important;color:#1d4ed8 !important;border-color:#bfdbfe !important}
.tag-seller{background:#f5f3ff !important;color:#6d28d9 !important;border-color:#ddd6fe !important}
.checklist-title{font-size:16px;font-weight:900;margin-bottom:2px}
.deal-table{width:100%;border-collapse:collapse;font-size:13px;background:#fff}
.deal-table th{background:#f8fafc;color:#475569;font-weight:900;padding:10px 8px;border:1px solid #eef2f7;text-align:center;white-space:nowrap}
.deal-table td{padding:9px 10px;border:1px solid #eef2f7}
.deal-table td.center{text-align:center;font-weight:800}
.deal-table td.group-cell{background:#ecfdf5;color:#047857;font-weight:900;text-align:center;vertical-align:middle}
.deal-table tr.total td{background:#eff6ff;color:#1e40af;font-weight:900}
.month-summary-table{font-size:16px;background:#fff}
.month-summary-table th{background:#f8fafc;color:#475569;font-weight:900;padding:14px 12px;border:1px solid #eef2f7;text-align:center;white-space:nowrap;font-size:15px}
.month-summary-table td{padding:14px 12px;border:1px solid #eef2f7;font-size:16px}
.month-summary-table td.center{text-align:center;font-weight:800}
@media(max-width:1000px){.field-grid{grid-template-columns:1fr}}
.gantt-wrap{overflow-x:auto;}
.gantt-grid{display:grid;grid-template-columns:repeat(7,minmax(110px,1fr));gap:5px 8px;align-items:center;}
.gantt-day-head{text-align:center;padding:8px 0;font-weight:900;color:#64748b;font-size:13px;}
.gantt-day-head .num{display:inline-block;width:32px;height:32px;line-height:32px;border-radius:50%;color:#111827;font-size:15px;margin-top:2px;}
.gantt-day-head .num.today{background:#2563eb;color:white;}
.gantt-bar{background:linear-gradient(90deg,#2563eb,#7c3aed);color:white;font-size:12px;font-weight:800;padding:7px 10px;border-radius:8px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;box-shadow:0 4px 10px rgba(37,99,235,.18);}
.gantt-bar.cancelled{background:#fecdd3;color:#9f1239;text-decoration:line-through;box-shadow:none;}
</style>
''', unsafe_allow_html=True)

def money(v:int)->str: return f"₩{v:,.0f}"

def pct(v):
    try: return f"{float(v)*100:.1f}%"
    except (TypeError, ValueError): return "-"

def fmt_money(v):
    try: return money(int(round(float(v))))
    except (TypeError, ValueError): return "-"

def fmt_text(v):
    if v is None: return "-"
    s = str(v).strip()
    return s if s else "-"

# ── 매출 집계 (스룩페이 주문서 → 일차별 집계) ─────────────────────────────
ORDER_REQUIRED_COLS = ['주문일','상품명','고객선택옵션','주문수량','취소수량','상품금액(옵션포함)','주문상태']

def parse_srookpay_orders(file_obj):
    """스룩페이 주문서 엑셀을 읽어 취소분을 반영한 정제 데이터프레임으로 변환."""
    df = pd.read_excel(file_obj)
    missing = [c for c in ORDER_REQUIRED_COLS if c not in df.columns]
    if missing:
        raise ValueError(f'주문서에서 다음 컬럼을 찾을 수 없습니다: {", ".join(missing)}')
    df = df[ORDER_REQUIRED_COLS].copy()
    df['주문일'] = pd.to_datetime(df['주문일'], errors='coerce')
    df = df.dropna(subset=['주문일'])
    for c in ['주문수량','취소수량','상품금액(옵션포함)']:
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
    df = df[~df['주문상태'].astype(str).str.contains('취소', na=False)]
    df['순주문수량'] = (df['주문수량'] - df['취소수량']).clip(lower=0)
    df = df[df['순주문수량'] > 0]
    qty_ratio = (df['순주문수량'] / df['주문수량'].replace(0, pd.NA)).fillna(1)
    df['순매출액'] = (df['상품금액(옵션포함)'] * qty_ratio).round(0)
    split = df['고객선택옵션'].astype(str).str.split(r'\s{2,}', n=1, regex=True)
    df['옵션그룹'] = split.str[0].str.strip()
    df['옵션상세'] = split.str[1].fillna('-').str.strip()
    df['주문일자'] = df['주문일'].dt.date
    return df

def build_deal_summary(df, start_date, deadline_date):
    """공구 시작일~마감일 기준으로 일차별/마감 버킷 집계 생성.
    마감일 당일 및 그 이후 주문은 모두 '마감' 컬럼으로 합산."""
    num_days = (deadline_date - start_date).days
    if num_days < 1:
        return None
    day_labels = [f'{i+1}일차' for i in range(num_days)]
    day_dates = [start_date + timedelta(days=i) for i in range(num_days)]

    df = df[df['주문일자'] >= start_date].copy()
    def bucket(d):
        if d >= deadline_date:
            return '마감'
        idx = (d - start_date).days
        return day_labels[idx] if 0 <= idx < num_days else None
    df['버킷'] = df['주문일자'].apply(bucket)
    df = df.dropna(subset=['버킷'])

    products = []
    for product_name, pdf in df.groupby('상품명', sort=False):
        groups = []
        for group_name, gdf in pdf.groupby('옵션그룹', sort=False):
            rows = []
            for opt_name, odf in gdf.groupby('옵션상세', sort=False):
                counts = {label: int(odf.loc[odf['버킷'] == label, '순주문수량'].sum()) for label in day_labels}
                counts['마감'] = int(odf.loc[odf['버킷'] == '마감', '순주문수량'].sum())
                rows.append({'option': opt_name, 'counts': counts, 'amount': float(odf['순매출액'].sum())})
            groups.append({'group': group_name, 'rows': rows})
        products.append({'product': product_name, 'groups': groups})

    return {
        'day_labels': day_labels, 'day_dates': day_dates, 'products': products,
        'total_qty': int(df['순주문수량'].sum()), 'total_amount': float(df['순매출액'].sum()),
    }

def render_deal_summary_html(summary):
    day_labels = summary['day_labels']
    html = '<div class="card" style="overflow-x:auto;padding:0;"><table class="deal-table"><thead><tr>'
    html += '<th>상품그룹</th><th>옵션</th>' + ''.join(f'<th>{l}</th>' for l in day_labels) + '<th>마감</th><th>상품 금액</th>'
    html += '</tr></thead><tbody>'
    for product in summary['products']:
        for group in product['groups']:
            rowspan = len(group['rows'])
            for i, row in enumerate(group['rows']):
                html += '<tr>'
                if i == 0:
                    html += f'<td class="group-cell" rowspan="{rowspan}">{group["group"]}</td>'
                html += f'<td>{row["option"]}</td>'
                html += ''.join(f'<td class="center">{row["counts"][l]}</td>' for l in day_labels)
                html += f'<td class="center">{row["counts"]["마감"]}</td>'
                html += f'<td class="center">{fmt_money(row["amount"])}</td></tr>'
    html += f'<tr class="total"><td colspan="2" style="text-align:center;">총 주문 수량 / 총 판매 금액</td>'
    html += ''.join('<td></td>' for _ in day_labels) + '<td></td>'
    html += f'<td class="center">{summary["total_qty"]}건 · {fmt_money(summary["total_amount"])}</td></tr>'
    html += '</tbody></table></div>'
    return html

def build_deal_summary_excel(summary, meta):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '매출집계'
    day_labels = summary['day_labels']
    headers = ['상품그룹', '옵션'] + day_labels + ['마감', '상품 금액']
    header_fill = PatternFill('solid', start_color='111827')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws['A1'] = f"공구 기간: {meta['start']} {meta['start_time']} ~ {meta['deadline']} {meta['deadline_time']}"
    ws['A1'].font = Font(bold=True, size=12)
    header_row = 3
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font, cell.fill, cell.border = header_font, header_fill, border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    r = header_row + 1
    for product in summary['products']:
        for group in product['groups']:
            start_row = r
            for row in group['rows']:
                ws.cell(row=r, column=2, value=row['option']).border = border
                for ci, label in enumerate(day_labels, start=3):
                    c = ws.cell(row=r, column=ci, value=row['counts'][label])
                    c.border, c.alignment = border, Alignment(horizontal='center')
                c = ws.cell(row=r, column=3 + len(day_labels), value=row['counts']['마감'])
                c.border, c.alignment = border, Alignment(horizontal='center')
                c = ws.cell(row=r, column=4 + len(day_labels), value=row['amount'])
                c.number_format, c.border = '#,##0', border
                r += 1
            end_row = r - 1
            gcell = ws.cell(row=start_row, column=1, value=group['group'])
            gcell.font = Font(bold=True)
            gcell.fill = PatternFill('solid', start_color='D9F2E3')
            gcell.alignment = Alignment(horizontal='center', vertical='center')
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            for rr in range(start_row, end_row + 1):
                ws.cell(row=rr, column=1).border = border

    r += 1
    ws.cell(row=r, column=1, value='총 주문 수량').font = Font(bold=True)
    ws.cell(row=r, column=2, value=summary['total_qty']).font = Font(bold=True, color='FF0000')
    r += 1
    ws.cell(row=r, column=1, value='총 판매 금액').font = Font(bold=True)
    amt_cell = ws.cell(row=r, column=2, value=summary['total_amount'])
    amt_cell.number_format, amt_cell.font = '#,##0', Font(bold=True, color='FF0000')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 26
    for i in range(len(day_labels) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(3 + i)].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def df_to_excel_bytes(df, sheet_name='Sheet1'):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    buf.seek(0)
    return buf

# ── 이벤트 추첨 / 개인정보 마스킹 ──────────────────────────────────────
NAME_COL_CANDIDATES = ['주문자명', '구매자명', '수취인명', '수령인명', '이름']
PHONE_COL_CANDIDATES = ['주문자연락처', '구매자연락처', '수취인연락처', '수령인연락처', '연락처']

def find_col(df, candidates):
    return next((c for c in candidates if c in df.columns), None)

def parse_order_contacts(file_obj):
    """추첨용: 주문서에서 이름/연락처만 추출 (취소 제외, 동일 연락처 중복 제거)."""
    df = pd.read_excel(file_obj)
    name_col = find_col(df, NAME_COL_CANDIDATES)
    phone_col = find_col(df, PHONE_COL_CANDIDATES)
    if not name_col or not phone_col:
        raise ValueError('이름/연락처 컬럼을 찾을 수 없습니다. (예: 주문자명·주문자연락처 또는 구매자명·구매자연락처)')
    out = df[[name_col, phone_col]].copy()
    out.columns = ['이름', '연락처']
    if '주문상태' in df.columns:
        out = out[~df['주문상태'].astype(str).str.contains('취소', na=False)]
    out = out.dropna(subset=['이름', '연락처'])
    out = out.drop_duplicates(subset=['연락처']).reset_index(drop=True)
    return out

def mask_name(name):
    s = str(name).strip()
    if len(s) <= 1: return s
    if len(s) == 2: return s[0] + '*'
    return s[0] + '*' * (len(s) - 2) + s[-1]

def mask_phone_last4(phone):
    digits = re.sub(r'\D', '', str(phone))
    return digits[-4:] if len(digits) >= 4 else digits

def build_masked_list(file_obj):
    df = pd.read_excel(file_obj)
    name_col = find_col(df, NAME_COL_CANDIDATES)
    phone_col = find_col(df, PHONE_COL_CANDIDATES)
    if not name_col or not phone_col:
        raise ValueError('이름/연락처 컬럼을 찾을 수 없습니다. (예: 구매자명·구매자연락처 또는 수취인명·수취인연락처)')
    out = pd.DataFrame()
    out['원본 이름'] = df[name_col]
    out['원본 연락처'] = df[phone_col]
    out['마스킹 이름'] = df[name_col].apply(mask_name)
    out['마스킹 연락처'] = df[phone_col].apply(mask_phone_last4)
    return out

# ── 대시보드 데이터 파이프라인 (공구현황판 / 결산자료 폴백 구조) ──────────────
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
BOARD_PATH = os.path.join(DATA_DIR, 'gonggu_board.xlsx')
SETTLEMENT_DIR = os.path.join(DATA_DIR, 'settlements')

def ensure_data_dirs():
    os.makedirs(SETTLEMENT_DIR, exist_ok=True)

def header_key(v):
    """헤더 셀 값에서 줄바꿈/공백 제거 후 비교용 키로 변환."""
    if v is None: return ''
    return re.sub(r'\s+', '', str(v))

def parse_year_month(text):
    """'2026.1월', '26년 5월' 같은 시트명에서 (연도,월)을 추출."""
    if not text: return None
    m = re.search(r'(\d{2,4})[.\s]*년?[.\s]*(\d{1,2})\s*월', str(text))
    if not m: return None
    y, mo = int(m.group(1)), int(m.group(2))
    if y < 100: y += 2000
    return (y, mo) if 1 <= mo <= 12 else None

def parse_board_sheet(ws):
    """공구현황판 월별 시트에서 '계' 합계 행 + 담당자별 실제 매출/GP 집계."""
    header_row = None
    for r in range(1, 20):
        for c in range(1, ws.max_column + 1):
            if header_key(ws.cell(row=r, column=c).value) == '셀러':
                header_row = r; break
        if header_row: break
    if header_row is None: return None
    col_map = {}
    for c in range(1, ws.max_column + 1):
        k = header_key(ws.cell(row=header_row, column=c).value)
        if k and k not in col_map: col_map[k] = c
    needed = ['셀러', '예상매출', '예상GP', '영업담당자', '매출', 'GP']
    if not all(k in col_map for k in needed): return None
    total_row = None
    for r in range(header_row + 1, min(header_row + 8, ws.max_row + 1)):
        for c in range(1, 5):
            if header_key(ws.cell(row=r, column=c).value) == '계':
                total_row = r; break
        if total_row: break
    if total_row is None: return None
    def val(row, key):
        v = ws.cell(row=row, column=col_map[key]).value
        return v if isinstance(v, (int, float)) else 0
    result = {
        'expected_revenue': val(total_row, '예상매출'), 'expected_gp': val(total_row, '예상GP'),
        'actual_revenue': val(total_row, '매출'), 'actual_gp': val(total_row, 'GP'), 'by_manager': {},
    }
    manager_col, rev_col, gp_col = col_map['영업담당자'], col_map['매출'], col_map['GP']
    for r in range(total_row + 1, ws.max_row + 1):
        manager = ws.cell(row=r, column=manager_col).value
        if not manager: continue
        rev = ws.cell(row=r, column=rev_col).value
        gp = ws.cell(row=r, column=gp_col).value
        rev = rev if isinstance(rev, (int, float)) else 0
        gp = gp if isinstance(gp, (int, float)) else 0
        if rev == 0 and gp == 0: continue
        key = str(manager).strip()
        d = result['by_manager'].setdefault(key, {'매출': 0, 'GP': 0})
        d['매출'] += rev; d['GP'] += gp
    return result

def parse_settlement_sheet(ws):
    """결산자료 시트에서 '영업팀 최종 합계' 행 + 담당자별(비고1) 매출/트리즈GP 집계."""
    header_row = None
    for r in range(1, 10):
        for c in range(1, ws.max_column + 1):
            if header_key(ws.cell(row=r, column=c).value) == '구분3':
                header_row = r; break
        if header_row: break
    if header_row is None: return None
    col_map = {}
    for c in range(1, ws.max_column + 1):
        k = header_key(ws.cell(row=header_row, column=c).value)
        if k and k not in col_map: col_map[k] = c
    needed = ['구분3', '매출', '트리즈GP', '비고1']
    if not all(k in col_map for k in needed): return None
    name_col, rev_col, gp_col, mgr_col = col_map['구분3'], col_map['매출'], col_map['트리즈GP'], col_map['비고1']
    total_row = None
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=name_col).value
        if v and '최종합계' in header_key(v):
            total_row = r; break
    if total_row is None: return None
    def numval(row, col):
        v = ws.cell(row=row, column=col).value
        return v if isinstance(v, (int, float)) else 0
    result = {'revenue': numval(total_row, rev_col), 'gp': numval(total_row, gp_col), 'by_manager': {}}
    for r in range(header_row + 1, total_row):
        manager = ws.cell(row=r, column=mgr_col).value
        if not manager: continue
        rev, gp = numval(r, rev_col), numval(r, gp_col)
        if rev == 0 and gp == 0: continue
        key = str(manager).strip()
        d = result['by_manager'].setdefault(key, {'매출': 0, 'GP': 0})
        d['매출'] += rev; d['GP'] += gp
    return result

def parse_board_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for sn in wb.sheetnames:
        ym = parse_year_month(sn)
        if not ym: continue
        parsed = parse_board_sheet(wb[sn])
        if parsed: out[ym] = parsed
    return out

def parse_settlement_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for sn in wb.sheetnames:
        ym = parse_year_month(sn)
        if not ym: continue
        parsed = parse_settlement_sheet(wb[sn])
        if parsed: out[ym] = parsed
    return out

def save_board_upload(uploaded_file):
    ensure_data_dirs()
    with open(BOARD_PATH, 'wb') as f:
        f.write(uploaded_file.getbuffer())

def save_settlement_upload(uploaded_file):
    ensure_data_dirs()
    raw = uploaded_file.getbuffer()
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ym = None
    for sn in wb.sheetnames:
        ym = parse_year_month(sn)
        if ym: break
    if ym is None:
        raise ValueError('시트명에서 연/월을 인식하지 못했습니다. (예: "26년 5월" 형식이어야 합니다)')
    with open(os.path.join(SETTLEMENT_DIR, f'{ym[0]}-{ym[1]:02d}.xlsx'), 'wb') as f:
        f.write(raw)
    return ym

def load_all_dashboard_data():
    ensure_data_dirs()
    board = parse_board_workbook(BOARD_PATH) if os.path.exists(BOARD_PATH) else {}
    settlements = {}
    for fn in os.listdir(SETTLEMENT_DIR):
        if fn.endswith('.xlsx'):
            settlements.update(parse_settlement_workbook(os.path.join(SETTLEMENT_DIR, fn)))
    return board, settlements

def merge_monthly(board, settlements):
    """월별로 결산 데이터가 있으면 결산을, 없으면 공구현황판(잠정)을 사용."""
    merged = {}
    for ym in sorted(set(board) | set(settlements)):
        if ym in settlements:
            s = settlements[ym]
            merged[ym] = {'revenue': s['revenue'], 'gp': s['gp'], 'by_manager': s['by_manager'], 'source': '결산 확정'}
        else:
            b = board[ym]
            merged[ym] = {'revenue': b['actual_revenue'], 'gp': b['actual_gp'], 'by_manager': b['by_manager'], 'source': '공구현황판 잠정'}
    return merged

def aggregate_by_manager(merged):
    totals = {}
    for ym, d in merged.items():
        for mgr, v in d['by_manager'].items():
            t = totals.setdefault(mgr, {'매출': 0, 'GP': 0})
            t['매출'] += v['매출']; t['GP'] += v['GP']
    return totals

def cumulative_by_manager_upto_month(board_data, year, upto_month):
    """해당 연도의 1월~upto_month월까지 담당자별 누적 매출/GP (공구현황판 기준)."""
    totals = {}
    for (y, m), d in board_data.items():
        if y == year and m <= upto_month:
            for mgr, v in d['by_manager'].items():
                t = totals.setdefault(mgr, {'매출': 0, 'GP': 0})
                t['매출'] += v['매출']; t['GP'] += v['GP']
    return totals
def filter_up_to_current_month(data_by_ym):
    """미래 월(아직 오지 않은 달) 시트는 대시보드 집계에서 제외."""
    today = date.today()
    cur = (today.year, today.month)
    return {ym: v for ym, v in data_by_ym.items() if ym <= cur}

def to_date_value(v):
    """엑셀 셀 값(datetime 또는 serial number)을 date로 변환."""
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, (int, float)):
        try: return from_excel(v).date()
        except Exception: return None
    return None

def format_deal_name(seller, product):
    """셀러/상품명을 합쳐서 표시용 이름을 만든다. 예: '초이끄'+'프레센티아/넥필' → '초이끄/프레센티아X넥필'
    (여러 '/' 중 마지막 하나만 콜라보 표시용 'X'로 바뀌고, 앞쪽 구분자는 '/' 그대로 유지)
    단, '(순한맛/오리지널맛)'처럼 괄호 안의 '/'는 '또는' 의미이므로 건드리지 않는다."""
    seller = str(seller).strip() if seller else ''
    product = str(product).strip() if product else ''
    parts = [p for p in (seller, product) if p]
    if not parts:
        return '(이름없음)'
    combined = '/'.join(parts)
    paren_spans = []
    def _stash(m):
        paren_spans.append(m.group(0))
        return f'\x00{len(paren_spans) - 1}\x00'
    protected = re.sub(r'\([^)]*\)', _stash, combined)
    matches = list(re.finditer(r'\s*/\s*', protected))
    if matches:
        last = matches[-1]
        protected = protected[:last.start()] + 'X' + protected[last.end():]
    return re.sub(r'\x00(\d+)\x00', lambda m: paren_spans[int(m.group(1))], protected)

def get_week_deal_list(board_path, year, month, week_start, week_end):
    """이번 주(월~일)에 시작하는 공구를 현황판 기재 순서 그대로 확정/취소로 분리.
    취소는 안내 문구(취소 시 예상매출·GP 삭제 후 숨김 처리)에 따라 '숨김 처리된 행'으로 판단."""
    if not os.path.exists(board_path): return None
    wb = openpyxl.load_workbook(board_path, data_only=True)
    sheet_name = next((sn for sn in wb.sheetnames if parse_year_month(sn) == (year, month)), None)
    if sheet_name is None: return None
    ws = wb[sheet_name]
    header_row, col_map = None, {}
    for r in range(1, 20):
        for c in range(1, ws.max_column + 1):
            if header_key(ws.cell(row=r, column=c).value) == '셀러':
                header_row = r; break
        if header_row: break
    if header_row is None: return None
    for c in range(1, ws.max_column + 1):
        k = header_key(ws.cell(row=header_row, column=c).value)
        if k and k not in col_map: col_map[k] = c
    seller_col, product_col = col_map.get('셀러'), col_map.get('진행상품')
    confirmed, cancelled = [], []
    for r in range(header_row + 1, ws.max_row + 1):
        seller = ws.cell(row=r, column=seller_col).value if seller_col else None
        start_d = to_date_value(ws.cell(row=r, column=3).value)  # C열=시작일
        end_d = to_date_value(ws.cell(row=r, column=5).value)    # E열=마감일
        if start_d is None or not (week_start <= start_d <= week_end): continue
        product = ws.cell(row=r, column=product_col).value if product_col else None
        date_label = f'{start_d.month}/{start_d.day}' if (not end_d or end_d == start_d) else f'{start_d.month}/{start_d.day}-{end_d.month}/{end_d.day}'
        entry = {'date_label': date_label, 'name': format_deal_name(seller, product)}
        if ws.row_dimensions[r].hidden:
            cancelled.append(entry)
        elif seller:
            confirmed.append(entry)
    return {'confirmed': confirmed, 'cancelled': cancelled}

def render_monthly_table_html(board_data):
    """연월/예상매출/예상GP/매출/GP/매출달성률/GP달성률 표. 달성률 100% 초과=파랑, 이하=빨강."""
    rows = ''
    for (y, m), d in sorted(board_data.items()):
        exp_rev, exp_gp = d['expected_revenue'] or 0, d['expected_gp'] or 0
        rev_pct = (d['actual_revenue'] / exp_rev * 100) if exp_rev else 0
        gp_pct = (d['actual_gp'] / exp_gp * 100) if exp_gp else 0
        rev_color = '#2563eb' if rev_pct > 100 else '#dc2626'
        gp_color = '#2563eb' if gp_pct > 100 else '#dc2626'
        rows += (f'<tr><td style="white-space:nowrap;"><b>{y % 100}년 {m}월</b></td>'
                 f'<td class="center">{money(exp_rev)}</td><td class="center">{money(exp_gp)}</td>'
                 f'<td class="center">{money(d["actual_revenue"])}</td><td class="center">{money(d["actual_gp"])}</td>'
                 f'<td class="center" style="color:{rev_color};font-weight:900;">{rev_pct:.1f}%</td>'
                 f'<td class="center" style="color:{gp_color};font-weight:900;">{gp_pct:.1f}%</td></tr>')
    return f'<div class="card" style="overflow-x:auto;padding:0;"><table class="month-summary-table" style="width:100%;border-collapse:collapse;"><thead><tr><th>연월</th><th>예상매출</th><th>예상GP</th><th>매출</th><th>GP</th><th>매출 달성률</th><th>GP 달성률</th></tr></thead><tbody>{rows}</tbody></table></div>'

def render_deal_list_html(title, items):
    if not items:
        body = '<div style="color:#94a3b8;font-size:13px;padding:6px 2px;">해당 없음</div>'
    else:
        body = ''.join(f'<div style="padding:7px 2px;border-bottom:1px solid #f1f5f9;font-size:13px;"><span style="color:#64748b;font-weight:800;margin-right:8px;">{it["date_label"]}</span>{it["name"]}</div>' for it in items)
    return f'<div class="card"><b>{title}</b><div style="margin-top:6px;">{body}</div></div>'

def parse_board_deals(board_path, year, month):
    """공구현황판 특정 월 시트의 개별 공구 행을 시트 기재 순서 그대로 리스트로 반환."""
    if not os.path.exists(board_path): return None
    wb = openpyxl.load_workbook(board_path, data_only=True)
    sheet_name = next((sn for sn in wb.sheetnames if parse_year_month(sn) == (year, month)), None)
    if sheet_name is None: return None
    ws = wb[sheet_name]
    header_row, col_map = None, {}
    for r in range(1, 20):
        for c in range(1, ws.max_column + 1):
            if header_key(ws.cell(row=r, column=c).value) == '셀러':
                header_row = r; break
        if header_row: break
    if header_row is None: return None
    for c in range(1, ws.max_column + 1):
        k = header_key(ws.cell(row=header_row, column=c).value)
        if k and k not in col_map: col_map[k] = c
    seller_col, product_col, manager_col = col_map.get('셀러'), col_map.get('진행상품'), col_map.get('영업담당자')
    deals = []
    for r in range(header_row + 1, ws.max_row + 1):
        seller = ws.cell(row=r, column=seller_col).value if seller_col else None
        start_d = to_date_value(ws.cell(row=r, column=3).value)
        end_d = to_date_value(ws.cell(row=r, column=5).value)
        is_cancelled = ws.row_dimensions[r].hidden
        if start_d is None or (not seller and not is_cancelled): continue
        product = ws.cell(row=r, column=product_col).value if product_col else None
        manager = ws.cell(row=r, column=manager_col).value if manager_col else None
        deals.append({
            'start': start_d, 'end': end_d or start_d,
            'seller': seller, 'product': product,
            'manager': str(manager).strip() if manager else '-',
            'status': '취소' if is_cancelled else '확정',
        })
    return deals

def get_deals_for_week(board_path, week_start, week_end):
    """주간 캘린더가 월 경계를 넘어갈 수 있으므로 관련된 월 시트를 모두 모아서 반환."""
    months_needed = {(week_start.year, week_start.month), (week_end.year, week_end.month)}
    all_deals = []
    any_sheet_found = False
    for (y, m) in months_needed:
        deals = parse_board_deals(board_path, y, m)
        if deals is not None:
            any_sheet_found = True
            all_deals.extend(deals)
    return all_deals if any_sheet_found else None

def parse_board_deals_all(board_path):
    """공구현황판의 모든 연/월 시트(25년/26년 등)를 통틀어 개별 공구 내역(매출/GP 포함)을 전부 수집."""
    if not os.path.exists(board_path): return []
    wb = openpyxl.load_workbook(board_path, data_only=True)
    all_deals = []
    for sn in wb.sheetnames:
        ym = parse_year_month(sn)
        if not ym: continue
        year, month = ym
        ws = wb[sn]
        header_row, col_map = None, {}
        for r in range(1, 20):
            for c in range(1, ws.max_column + 1):
                if header_key(ws.cell(row=r, column=c).value) == '셀러':
                    header_row = r; break
            if header_row: break
        if header_row is None: continue
        for c in range(1, ws.max_column + 1):
            k = header_key(ws.cell(row=header_row, column=c).value)
            if k and k not in col_map: col_map[k] = c
        seller_col, product_col = col_map.get('셀러'), col_map.get('진행상품')
        manager_col, rev_col, gp_col = col_map.get('영업담당자'), col_map.get('매출'), col_map.get('GP')
        for r in range(header_row + 1, ws.max_row + 1):
            seller = ws.cell(row=r, column=seller_col).value if seller_col else None
            start_d = to_date_value(ws.cell(row=r, column=3).value)
            end_d = to_date_value(ws.cell(row=r, column=5).value)
            is_cancelled = ws.row_dimensions[r].hidden
            if start_d is None or (not seller and not is_cancelled): continue
            product = ws.cell(row=r, column=product_col).value if product_col else None
            manager = ws.cell(row=r, column=manager_col).value if manager_col else None
            rev = ws.cell(row=r, column=rev_col).value if rev_col else None
            gp = ws.cell(row=r, column=gp_col).value if gp_col else None
            all_deals.append({
                'year': year, 'month': month,
                'start': start_d, 'end': end_d or start_d,
                'seller': seller, 'product': product,
                'manager': str(manager).strip() if manager else '-',
                'status': '취소' if is_cancelled else '확정',
                'revenue': rev if isinstance(rev, (int, float)) else 0,
                'gp': gp if isinstance(gp, (int, float)) else 0,
            })
    return all_deals

def search_deal_history(all_deals, seller_query, product_query):
    seller_q = (seller_query or '').strip()
    product_q = (product_query or '').strip()
    def match(d):
        if seller_q and seller_q not in str(d['seller'] or ''): return False
        if product_q and product_q not in str(d['product'] or ''): return False
        return True
    results = [d for d in all_deals if match(d)]
    results.sort(key=lambda d: (d['year'], d['month'], d['start']), reverse=True)
    return results

# ── 공구 알람 텍스트 생성 (영업매니저 작성 체크리스트 → 공유용 텍스트) ──────
WEEKDAY_KR = ['월','화','수','목','금','토','일']

def fmt_date_kr(d):
    return f'{d.month:02d}/{d.day:02d}({WEEKDAY_KR[d.weekday()]})'

def parse_schedule_from_sheet_name(sheet_name, year):
    """시트명 앞부분의 'MMDD-MMDD' 또는 'MMDD-DD'(같은 달) 패턴에서 공구 일정을 추출."""
    m = re.match(r'^(\d{2})(\d{2})-(\d{2})(\d{2})', sheet_name)
    if m:
        sm, sd, em, ed = map(int, m.groups())
        try: return date(year, sm, sd), date(year, em, ed)
        except ValueError: return None, None
    m = re.match(r'^(\d{2})(\d{2})-(\d{1,2})\b', sheet_name)
    if m:
        sm, sd, ed = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try: return date(year, sm, sd), date(year, sm, ed)
        except ValueError: return None, None
    return None, None

def format_shipping_text(raw):
    if not raw: return ''
    raw_lines = [l.strip() for l in str(raw).replace('▶', '-').split('\n') if l.strip()]
    lines = [re.sub(r'^-(?!\s)', '- ', l) for l in raw_lines]
    out, prev_bullet = [], None
    for line in lines:
        is_bullet = line.startswith('-')
        if prev_bullet is False and is_bullet:
            out.append('')
        out.append(line)
        prev_bullet = is_bullet
    return '\n'.join(out)

def format_event_text(raw):
    if not raw: return ''
    lines = [l.strip() for l in str(raw).split('\n')]
    content = [l for l in lines if l and not re.fullmatch(r'\[[^\]]*\]', l)]
    return '\n'.join(f'{i+1}. {l}' for i, l in enumerate(content)) if content else ''

def find_label_row(ws, keywords, col=2, max_row=30):
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v and all(kw in str(v) for kw in keywords):
            return r
    return None

def build_gonggu_alarm_text(ws, year):
    def val(keywords):
        r = find_label_row(ws, keywords)
        return ws.cell(row=r, column=3).value if r else None

    seller = val(['셀러']) or ''
    product = val(['진행', '상품']) or ''
    open_text = val(['결제창', '오픈']) or ''
    close_text = val(['결제창', '마감', '시간']) or ''
    link = val(['결제창', '링크']) or ''
    same_day = val(['당일', '발송', '마감']) or ''
    event_raw = val(['이벤트'])

    header_row, ship_col = None, None
    for r in range(1, 20):
        for c in range(1, ws.max_column + 1):
            if '배송비' in header_key(ws.cell(row=r, column=c).value or ''):
                header_row, ship_col = r, c; break
        if header_row: break
    shipping_raw = ws.cell(row=header_row + 1, column=ship_col).value if header_row else ''

    start_d, end_d = parse_schedule_from_sheet_name(ws.title, year)
    schedule_line = f'{fmt_date_kr(start_d)}-{fmt_date_kr(end_d)}' if start_d else '(시트명에서 일정을 인식하지 못했습니다 — 직접 확인해주세요)'

    parts = [
        f'⭐{seller} X {product}⭐', '',
        f'➡️ 일정: {schedule_line}',
        f'- 결제창 오픈 : {open_text}',
        f'- 결제창 마감 : {close_text}', '',
        '➡️ 배송',
        format_shipping_text(shipping_raw), '',
        '➡️ 당일 배송 시간',
        str(same_day), '',
        '➡️ 결제창 링크',
        str(link), '',
        '➡️ 이벤트',
        format_event_text(event_raw) or '(없음)',
    ]
    return '\n'.join(parts)
def render_gantt_calendar_html(week_start, week_end, deals):
    """요일 헤더 + 공구별 연속 막대(간트차트 스타일) 캘린더."""
    day_names = ['일', '월', '화', '수', '목', '금', '토']
    today = date.today()
    header = ''
    for i in range(7):
        d = week_start + timedelta(days=i)
        num_class = 'num today' if d == today else 'num'
        header += f'<div class="gantt-day-head" style="grid-column:{i+1};grid-row:1;">{day_names[i]}<br><span class="{num_class}">{d.day}</span></div>'
    bars, row = '', 2
    for deal in sorted(deals, key=lambda x: (x['start'], str(x['seller']))):
        if deal['end'] < week_start or deal['start'] > week_end: continue
        start_off = max(0, (deal['start'] - week_start).days)
        end_off = min(6, (deal['end'] - week_start).days)
        klass = 'gantt-bar cancelled' if deal['status'] == '취소' else 'gantt-bar'
        name = format_deal_name(deal['seller'], deal['product'])
        bars += f'<div class="{klass}" style="grid-column:{start_off+1} / {end_off+2};grid-row:{row};" title="{name}">{name}</div>'
        row += 1
    return f'<div class="card gantt-wrap"><div class="gantt-grid">{header}{bars}</div></div>'

manager_df=pd.DataFrame([{"담당자":"매니저 A","공구수":8,"매출":230_000_000,"GP":18_600_000,"KPI":250_000_000},{"담당자":"매니저 B","공구수":6,"매출":180_000_000,"GP":13_500_000,"KPI":220_000_000},{"담당자":"매니저 C","공구수":4,"매출":120_000_000,"GP":8_640_000,"KPI":180_000_000}])
manager_df['달성률']=(manager_df['매출']/manager_df['KPI']*100).round(1)

with st.sidebar:
    st.markdown('## 📊 TRIZ 영업실 업무 프로그램')
    st.caption('보고용 PC 프로그램 데모')
    page=st.radio('메뉴',['🏠 메인 대시보드','📅 공구 일정','💰 매출 집계','🎁 이벤트 추첨','👩 담당자별 매출','🔍 히스토리 검색','📢 공구 알람'],label_visibility='collapsed')
    st.markdown('---')
    st.caption('※ 현재 버전은 보고용 시안입니다. 일부 업로드/버튼은 화면 시연용입니다.')

st.markdown('<div class="main-title">TRIZ 영업실 업무 프로그램</div>',unsafe_allow_html=True)
st.markdown('<div class="subtle">주문서·매출 파일 업로드 기반으로 영업실 루틴 업무를 자동화하는 내부 포털 PC 데모</div>',unsafe_allow_html=True)

if page=='🏠 메인 대시보드':
    st.markdown('<div class="section-title">🏠 메인 대시보드</div>',unsafe_allow_html=True)
    st.markdown('<div class="help">공구현황판을 업로드하면 해당 월의 매출·GP와 예상 대비 실적이 자동으로 표시됩니다. 업로드한 파일은 앱 폴더(data/)에 계속 저장됩니다.</div>',unsafe_allow_html=True)

    with st.expander('📁 공구현황판 업로드',expanded=False):
        board_up=st.file_uploader('공구현황판 업로드 (여러 월 시트가 포함된 워크북 1개)',type=['xlsx'],key='board_upload')
        if board_up is not None:
            save_board_upload(board_up)
            st.success('공구현황판 저장 완료 (다음부터는 자동으로 반영됩니다)')

    board_data_all=load_all_dashboard_data()[0]
    board_data=filter_up_to_current_month(board_data_all)

    if os.path.exists(BOARD_PATH):
        updated_dt=datetime.fromtimestamp(os.path.getmtime(BOARD_PATH))
        st.markdown(f'<span class="chip active">🕒 업데이트 : {updated_dt.strftime("%Y-%m-%d %H:%M")}</span>',unsafe_allow_html=True)

    if not board_data:
        st.markdown('<div class="dark-card"><div style="display:flex;justify-content:space-between;gap:20px;align-items:center;"><div><b style="font-size:18px;">아직 업로드된 공구현황판이 없습니다</b><br><span>위 업로드 영역에서 공구현황판을 올리면 아래 지표가 실제 데이터로 자동 전환됩니다. (지금은 샘플 수치)</span></div></div></div>',unsafe_allow_html=True)
        c1,c2,c3,c4=st.columns(4); c1.metric('이번 달 총매출(샘플)','₩772,000,000'); c2.metric('이번 달 GP(샘플)','₩61,200,000','GP율 7.92%'); c3.metric('예상매출 대비(샘플)','104.3%','+₩12,400,000'); c4.metric('예상GP 대비(샘플)','98.1%','-₩1,200,000')
    else:
        latest_ym=max(board_data.keys())
        cur=board_data[latest_ym]
        exp_rev,exp_gp=cur['expected_revenue'] or 0,cur['expected_gp'] or 0
        rev_pct=(cur['actual_revenue']/exp_rev*100) if exp_rev else 0
        gp_pct=(cur['actual_gp']/exp_gp*100) if exp_gp else 0
        rev_diff=cur['actual_revenue']-exp_rev
        gp_diff=cur['actual_gp']-exp_gp
        c1,c2,c3,c4=st.columns(4)
        c1.metric(f'{latest_ym[0]}년 {latest_ym[1]}월 매출',money(cur['actual_revenue']))
        c2.metric('GP',money(cur['actual_gp']))
        c3.metric('예상매출 대비',f'{rev_pct:.1f}%',money(rev_diff))
        c4.metric('예상GP 대비',f'{gp_pct:.1f}%',money(gp_diff))

        st.markdown('<div class="section-title" style="font-size:1.15rem;">📊 월별 매출/GP 현황</div>',unsafe_allow_html=True)
        st.markdown(render_monthly_table_html(board_data),unsafe_allow_html=True)

    st.markdown('<div class="section-title">📌 주간 요약</div>',unsafe_allow_html=True)
    today=date.today()
    week_start=today-timedelta(days=today.weekday())
    week_end=week_start+timedelta(days=6)
    week_list=get_week_deal_list(BOARD_PATH,today.year,today.month,week_start,week_end)
    if week_list is None:
        st.markdown('<div class="card"><b>이번주 공구 현황</b><br><br><span class="chip">공구현황판을 업로드하면 표시됩니다</span></div>',unsafe_allow_html=True)
    else:
        st.markdown(f'<span class="chip">{week_start.month}/{week_start.day}~{week_end.month}/{week_end.day}</span>',unsafe_allow_html=True)
        wc1,wc2=st.columns(2)
        with wc1: st.markdown(render_deal_list_html(f'✅ 확정 공구 ({len(week_list["confirmed"])}건)',week_list['confirmed']),unsafe_allow_html=True)
        with wc2: st.markdown(render_deal_list_html(f'❌ 취소 공구 ({len(week_list["cancelled"])}건)',week_list['cancelled']),unsafe_allow_html=True)

elif page=='📅 공구 일정':
    st.markdown('<div class="section-title">📅 공구 일정</div>',unsafe_allow_html=True)
    st.markdown('<div class="help">공구현황판을 업로드하면 주간 캘린더가 자동으로 반영됩니다. ("🏠 메인 대시보드"에서 업로드)</div>',unsafe_allow_html=True)

    st.session_state.setdefault('cal_week_offset',0)
    today=date.today()
    base_sunday=today-timedelta(days=(today.weekday()+1)%7)

    nav1,nav2,nav3,nav4=st.columns([1,1,3,1])
    with nav1:
        if st.button('◀ 이전 주',use_container_width=True): st.session_state['cal_week_offset']-=1
    with nav2:
        if st.button('오늘',use_container_width=True): st.session_state['cal_week_offset']=0
    with nav4:
        if st.button('다음 주 ▶',use_container_width=True): st.session_state['cal_week_offset']+=1

    week_start=base_sunday+timedelta(weeks=st.session_state['cal_week_offset'])
    week_end=week_start+timedelta(days=6)
    with nav3:
        st.markdown(f'<div style="text-align:center;font-weight:900;padding-top:8px;">{week_start.year}년 {week_start.month}월 {week_start.day}일 ~ {week_end.month}월 {week_end.day}일</div>',unsafe_allow_html=True)

    all_deals=get_deals_for_week(BOARD_PATH,week_start,week_end)

    if all_deals is None:
        st.info('해당 기간의 공구현황판 데이터가 없습니다. "🏠 메인 대시보드"에서 업로드해주세요.')
    else:
        managers=sorted({d['manager'] for d in all_deals if d['manager'] and d['manager']!='-'})
        _,_,f1,f2,f3=st.columns([.8,.8,.8,.8,1.2])
        with f1: status_confirmed=st.checkbox('확정',True)
        with f2: status_cancelled=st.checkbox('취소',True)
        with f3: manager_filter=st.selectbox('담당자 필터',['전체']+managers)

        def keep(d):
            if d['status']=='확정' and not status_confirmed: return False
            if d['status']=='취소' and not status_cancelled: return False
            if manager_filter!='전체' and d['manager']!=manager_filter: return False
            return True
        filtered=[d for d in all_deals if keep(d)]

        st.markdown(render_gantt_calendar_html(week_start,week_end,filtered),unsafe_allow_html=True)

elif page=='💰 매출 집계':
    st.markdown('<div class="section-title">💰 매출 집계</div>',unsafe_allow_html=True)
    st.markdown('<div class="help">스룩페이 주문서 엑셀을 업로드하면 일차별 주문건수와 총 매출을 자동 집계합니다. 마감일에는 해당일 주문 + 그 이후 주문이 모두 포함됩니다.</div>',unsafe_allow_html=True)
    s1,s2,s3,s4=st.columns(4)
    with s1: deal_start=st.date_input('공구 시작일',value=date.today(),key='deal_start')
    with s2: deal_start_time=st.time_input('공구 시작 시간',value=time(10,0),key='deal_start_time')
    with s3: deal_end=st.date_input('공구 마감일',value=date.today()+timedelta(days=3),key='deal_end')
    with s4: deal_end_time=st.time_input('공구 마감 시간',value=time(23,59),key='deal_end_time')

    order_file=st.file_uploader('📤 스룩페이 주문서 업로드 (.xlsx)',type=['xlsx'],key='deal_order_upload')

    if deal_end <= deal_start:
        st.warning('마감일은 시작일보다 뒤여야 합니다.')
    elif order_file is not None:
        try:
            order_df=parse_srookpay_orders(order_file)
        except ValueError as e:
            order_df=None
            st.error(str(e))
        if order_df is not None:
            if order_df.empty:
                st.warning('유효한(취소되지 않은) 주문 건이 없습니다.')
            else:
                summary=build_deal_summary(order_df, deal_start, deal_end)
                if summary is None or not summary['products']:
                    st.warning('선택하신 공구 기간(시작일~마감일)에 해당하는 주문이 없습니다. 날짜를 확인해주세요.')
                else:
                    left,right=st.columns([3,1])
                    with left:
                        chips=f'<span class="chip active">트리즈창</span><span class="chip">시작 {deal_start.month}/{deal_start.day} {deal_start_time.strftime("%H:%M")}</span><span class="chip">마감 {deal_end.month}/{deal_end.day} {deal_end_time.strftime("%H:%M")}</span>'
                        st.markdown(f'<div class="card">{chips}{render_deal_summary_html(summary)}</div>',unsafe_allow_html=True)
                    with right:
                        st.markdown(f'<div class="dark-card"><span>총 매출</span><div style="font-size:26px;font-weight:900;letter-spacing:-.04em;margin:10px 0;">{fmt_money(summary["total_amount"])}</div><span>총 주문수량</span><div style="font-size:24px;font-weight:900;margin:8px 0 4px;">{summary["total_qty"]}건</div></div>',unsafe_allow_html=True)
                        meta={'start':deal_start.strftime('%Y-%m-%d'),'start_time':deal_start_time.strftime('%H:%M'),'deadline':deal_end.strftime('%Y-%m-%d'),'deadline_time':deal_end_time.strftime('%H:%M')}
                        excel_buf=build_deal_summary_excel(summary, meta)
                        st.download_button('📥 엑셀 다운로드',data=excel_buf,file_name=f'매출집계_{deal_start}~{deal_end}.xlsx',mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',use_container_width=True)
    else:
        st.info('공구 시작일/마감일을 확인한 뒤 스룩페이 주문서를 업로드하면 자동으로 집계됩니다.')



elif page=='🎁 이벤트 추첨':
    st.markdown('<div class="section-title">🎁 이벤트 추첨</div>',unsafe_allow_html=True)
    tab1,tab2=st.tabs(['주문서 기반 당첨자 추첨','당첨자 개인정보 마스킹'])
    with tab1:
        st.markdown('<div class="help">스룩페이 주문서를 업로드하고 당첨 인원을 정하면 랜덤으로 당첨자를 추첨합니다. (취소 주문 제외, 동일 연락처는 1명으로 처리)</div>',unsafe_allow_html=True)
        a,b,c=st.columns([2,1,1])
        with a: raffle_file=st.file_uploader('주문서 엑셀 업로드',type=['xlsx'],key='raffle_order')
        with b: winner_count=st.number_input('당첨 인원',min_value=1,value=5,key='winner_count')
        with c:
            st.markdown('<div style="height:28px;"></div>',unsafe_allow_html=True)
            draw=st.button('🎲 랜덤 추첨',type='primary',use_container_width=True)

        if draw:
            if raffle_file is None:
                st.warning('주문서 엑셀을 먼저 업로드해주세요.')
            else:
                try:
                    pool=parse_order_contacts(raffle_file)
                except ValueError as e:
                    pool=None
                    st.error(str(e))
                if pool is not None:
                    if pool.empty:
                        st.warning('추첨 가능한 주문 건이 없습니다.')
                    else:
                        n=min(int(winner_count), len(pool))
                        winners=pool.sample(n=n).reset_index(drop=True)
                        winners.insert(0,'순번',range(1,len(winners)+1))
                        winners['상태']='당첨'
                        st.session_state['raffle_winners']=winners
                        if n < winner_count:
                            st.info(f'추첨 대상이 {len(pool)}명뿐이라 {n}명만 추첨되었습니다.')

        if 'raffle_winners' in st.session_state:
            winners=st.session_state['raffle_winners']
            st.markdown(f'<div class="section-title" style="font-size:1.1rem;">🎉 당첨자 명단 ({len(winners)}명)</div>',unsafe_allow_html=True)
            st.dataframe(winners,use_container_width=True,hide_index=True)
            st.download_button('📥 당첨자 명단 엑셀 다운로드',data=df_to_excel_bytes(winners,'당첨자명단'),file_name='이벤트_당첨자_명단.xlsx',mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',key='raffle_download')

    with tab2:
        st.markdown('<div class="help">당첨자 명단 엑셀을 업로드하면 이름은 가운데 글자를, 연락처는 뒤 4자리만 남기고 자동으로 마스킹합니다.</div>',unsafe_allow_html=True)
        mask_file=st.file_uploader('당첨자 엑셀 업로드',type=['xlsx'],key='masking')
        if mask_file is not None:
            try:
                masked=build_masked_list(mask_file)
            except ValueError as e:
                masked=None
                st.error(str(e))
            if masked is not None:
                st.dataframe(masked,use_container_width=True,hide_index=True)
                st.download_button('📥 마스킹 결과 엑셀 다운로드',data=df_to_excel_bytes(masked,'마스킹결과'),file_name='당첨자_마스킹.xlsx',mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',key='mask_download')

elif page=='👩 담당자별 매출':
    st.markdown('<div class="section-title">👩 담당자별 매출 집계</div>',unsafe_allow_html=True)
    board_data=filter_up_to_current_month(load_all_dashboard_data()[0])

    if not board_data:
        st.info('공구현황판이 아직 없습니다. "🏠 메인 대시보드"에서 업로드하면 실제 담당자별 데이터로 전환됩니다. (지금은 샘플 수치)')
        st.markdown('<div class="card">',unsafe_allow_html=True)
        for _,row in manager_df.iterrows():
            st.markdown(f'<div style="display:grid;grid-template-columns:140px 1fr 180px;gap:16px;align-items:center;padding:14px 0;border-bottom:1px solid #eef2f7;"><div><b>{row["담당자"]}</b><br><span style="color:#64748b;font-size:12px;">공구 {int(row["공구수"])}건</span></div><div><div class="bar-bg"><div class="bar-fill" style="width:{min(row["달성률"],100)}%;"></div></div></div><div style="text-align:right;"><b>{money(int(row["매출"]))}</b><br><span style="color:#64748b;font-size:12px;">GP {money(int(row["GP"]))}</span></div></div>',unsafe_allow_html=True)
        st.markdown('</div>',unsafe_allow_html=True)
    else:
        years=sorted({y for (y,m) in board_data})
        cur_year=years[-1] if years else date.today().year
        st.markdown(f'<div class="help">{cur_year}년 누적 매출/GP (공구현황판 기준, 업로드된 월 전체 합산)</div>',unsafe_allow_html=True)
        totals=aggregate_by_manager(board_data)
        real_df=pd.DataFrame([{'담당자':m,'매출':v['매출'],'GP':v['GP']} for m,v in totals.items()])
        real_df=real_df.sort_values('매출',ascending=False).reset_index(drop=True)
        max_rev=real_df['매출'].max() if len(real_df) else 1

        st.markdown('<div class="card">',unsafe_allow_html=True)
        for _,row in real_df.iterrows():
            bar_pct=(row['매출']/max_rev*100) if max_rev else 0
            st.markdown(f'<div style="display:grid;grid-template-columns:140px 1fr 180px;gap:16px;align-items:center;padding:14px 0;border-bottom:1px solid #eef2f7;"><div><b>{row["담당자"]}</b></div><div><div class="bar-bg"><div class="bar-fill" style="width:{bar_pct:.1f}%;"></div></div></div><div style="text-align:right;"><b>매출 {money(row["매출"])}</b><br><b>GP {money(row["GP"])}</b></div></div>',unsafe_allow_html=True)
        st.markdown('</div>',unsafe_allow_html=True)

        st.markdown('<div class="section-title" style="font-size:1.1rem;">📆 담당자별(월별) 누적 조회</div>',unsafe_allow_html=True)
        managers=real_df['담당자'].tolist()
        months_avail=sorted({m for (y,m) in board_data if y==cur_year})
        sel1,sel2=st.columns(2)
        with sel1: manager_sel=st.selectbox('담당자',['전체']+managers)
        with sel2: month_sel=st.selectbox(f'{cur_year}년 조회월',months_avail,index=len(months_avail)-1,format_func=lambda m:f'{m}월') if months_avail else None

        if month_sel:
            cum=cumulative_by_manager_upto_month(board_data,cur_year,month_sel)
            if manager_sel!='전체':
                cum={manager_sel:cum.get(manager_sel,{'매출':0,'GP':0})}
            else:
                cum={m:cum.get(m,{'매출':0,'GP':0}) for m in managers}
            cum_df=pd.DataFrame([{'담당자':m,'누적 매출':money(v['매출']),'누적 GP':money(v['GP'])} for m,v in cum.items()])
            st.dataframe(cum_df,use_container_width=True,hide_index=True)

elif page=='🔍 히스토리 검색':
    st.markdown('<div class="section-title">🔍 셀러/상품 히스토리 검색</div>',unsafe_allow_html=True)
    st.markdown('<div class="help">공구현황판에 있는 모든 연/월 시트(25년·26년 등)를 통틀어 검색합니다. 셀러명 또는 상품명 일부만 입력해도 검색됩니다.</div>',unsafe_allow_html=True)

    if not os.path.exists(BOARD_PATH):
        st.info('공구현황판이 아직 없습니다. "🏠 메인 대시보드"에서 업로드해주세요.')
    else:
        s1,s2=st.columns(2)
        with s1: seller_q=st.text_input('셀러명 검색',placeholder='예: 초이끄')
        with s2: product_q=st.text_input('상품명 검색',placeholder='예: 프레센티아')

        if seller_q or product_q:
            all_deals=parse_board_deals_all(BOARD_PATH)
            results=search_deal_history(all_deals,seller_q,product_q)
            if not results:
                st.warning('검색 결과가 없습니다.')
            else:
                st.markdown(f'<div class="help">총 {len(results)}건 검색됨 (최신순)</div>',unsafe_allow_html=True)
                rows=[]
                for d in results:
                    date_label=f'{d["start"].year}.{d["start"].month}/{d["start"].day}' if d['start']==d['end'] else f'{d["start"].year}.{d["start"].month}/{d["start"].day}-{d["end"].month}/{d["end"].day}'
                    rows.append({'진행시기':date_label,'셀러X상품':format_deal_name(d['seller'],d['product']),'담당자':d['manager'],'상태':d['status'],'매출':money(d['revenue']),'GP':money(d['gp'])})
                st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
        else:
            st.info('셀러명 또는 상품명을 입력하면 히스토리가 표시됩니다.')

elif page=='📢 공구 알람':
    st.markdown('<div class="section-title">📢 공구 알람 생성</div>',unsafe_allow_html=True)
    st.markdown('<div class="help">영업매니저가 작성한 "공동구매 진행 전 체크리스트" 엑셀을 업로드하면 공유용 공구 알람 텍스트를 자동으로 만들어드립니다.</div>',unsafe_allow_html=True)

    alarm_file=st.file_uploader('체크리스트 엑셀 업로드 (.xlsx)',type=['xlsx'],key='alarm_upload')
    alarm_year=date.today().year

    if alarm_file is not None:
        try:
            wb=openpyxl.load_workbook(alarm_file,data_only=True)
        except Exception as e:
            wb=None
            st.error(f'파일을 읽는 중 문제가 발생했습니다: {e}')
        if wb is not None:
            sheet_name=st.selectbox('상품(시트) 선택',wb.sheetnames) if len(wb.sheetnames)>1 else wb.sheetnames[0]
            ws=wb[sheet_name]
            alarm_text=build_gonggu_alarm_text(ws,alarm_year)
            st.text_area('📋 공구 알람 (전체 선택해서 복사하세요)',value=alarm_text,height=440)
            st.download_button('📥 텍스트 파일로 다운로드',data=alarm_text.encode('utf-8'),file_name=f'공구알람_{sheet_name}.txt',mime='text/plain')
            st.caption('※ "일정" 날짜는 시트 이름(예: "0713-15 ...")에서 자동으로 인식하며, 연도는 현재 연도로 자동 적용됩니다. 예전 공구의 경우 요일이 다르게 나올 수 있어요.')
